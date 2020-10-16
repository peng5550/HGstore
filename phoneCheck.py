import threading
import json
from mttkinter import mtTkinter as mtk
from tkinter import ttk, Scrollbar, filedialog
from tkinter.messagebox import showinfo, showerror
import asyncio
import aiohttp
from openpyxl import Workbook, load_workbook
import re


class Application(object):

    def __init__(self):
        self.semNum = 5
        self.treeIndex = 1
        self.totals = 0
        self.totalData = [["手机号", "是否有优惠"]]
        self.__createGUI()

    def __createGUI(self):
        self.root = mtk.Tk()
        self.root.title("移动惠购用户查询工具")
        self.root.geometry("720x600")

        self.settings = mtk.LabelFrame(text="设置", fg="blue")
        self.settings.place(x=50, y=20, width=600, height=130)

        self.shopId = mtk.Label(self.settings, text="店铺  ID：")
        self.shopId.place(x=30, y=10, width=60, height=20)
        self.shopIdText = mtk.Entry(self.settings)
        self.shopIdText.place(x=85, y=10, width=80, height=20)

        self.workNo = mtk.Label(self.settings, text="工   号：")
        self.workNo.place(x=200, y=10, width=60, height=20)
        self.workNoText = mtk.Entry(self.settings)
        self.workNoText.place(x=260, y=10, width=80, height=20)

        self.semaphoreNum = mtk.Label(self.settings, text="并发数量：")
        self.semaphoreNum.place(x=30, y=70, width=60, height=20)
        self.entrySemaphoreNum = mtk.Entry(self.settings)
        self.entrySemaphoreNum.place(x=85, y=70, width=80, height=20)

        self.timeSleep = mtk.Label(self.settings, text="延迟时间：")
        self.timeSleep.place(x=200, y=70, width=60, height=20)
        self.entrySleep = mtk.Entry(self.settings)
        self.entrySleep.place(x=260, y=70, width=80, height=20)

        self.statusNow = mtk.Label(self.settings, text="当前进度：")
        self.statusNow.place(x=420, y=10, width=60, height=20)
        self.statusNowText = mtk.Label(self.settings, text="0/0")
        self.statusNowText.place(x=450, y=50, width=60, height=20)

        self.showDataBox = mtk.LabelFrame(self.root, text="数据信息", fg="blue")
        self.showDataBox.place(x=50, y=170, width=400, height=400)
        title = ['1', '2', '3']
        self.box = ttk.Treeview(self.showDataBox, columns=title, show='headings')
        self.box.place(x=20, y=15, width=360, height=340)
        self.box.column('1', width=50, anchor='center')
        self.box.column('2', width=200, anchor='center')
        self.box.column('3', width=100, anchor='center')
        self.box.heading('1', text='序号')
        self.box.heading('2', text='手机号')
        self.box.heading('3', text='是否有优惠')
        self.VScroll1 = Scrollbar(self.box, orient='vertical', command=self.box.yview)
        self.VScroll1.pack(side="right", fill="y")
        self.box.configure(yscrollcommand=self.VScroll1.set)

        self.btnBox = mtk.LabelFrame(self.root, text="任务栏", fg="blue")
        self.btnBox.place(x=480, y=170, width=170, height=400)

        self.fileBox = mtk.LabelFrame(self.btnBox)
        self.fileBox.place(x=15, y=30, width=140, height=120)

        self.loadExcel = mtk.Button(self.fileBox, text="导入Excel", command=lambda: self.thread_it(self.__loadExcel))
        self.loadExcel.place(x=15, y=20, width=100, height=30)
        self.btnEnd = mtk.Button(self.fileBox, text="导出Excel", command=lambda: self.thread_it(self.__saveExcel))
        self.btnEnd.place(x=15, y=70, width=100, height=30)

        self.btnBbox = mtk.LabelFrame(self.btnBox)
        self.btnBbox.place(x=15, y=200, width=140, height=120)
        self.btnStart = mtk.Button(self.btnBbox, text="开始", command=lambda: self.thread_it(self.start))
        self.btnStart.place(x=15, y=20, width=100, height=30)
        self.btnStop = mtk.Button(self.btnBbox, text="停止", command=lambda: self.thread_it(self.stop))
        self.btnStop.place(x=15, y=70, width=100, height=30)

    def deleteTree(self):
        x = self.box.get_children()
        for item in x:
            self.box.delete(item)

    def semaphoreNumSettings(self):
        semNum = self.entrySemaphoreNum.get().strip()
        if semNum:
            try:
                self.semNum = int(semNum)
            except:
                showerror("错误信息", "输入错误!")
        else:
            showerror("错误信息", "请选择并发数量!")

    def __loadExcel(self):
        self.treeIndex = 1
        excelPath = filedialog.askopenfilename(title=u'选择文件')
        if excelPath:
            try:
                self.excelData = []
                self.totalData = [["手机号", "是否有优惠"]]
                self.deleteTree()
                wb = load_workbook(excelPath)
                ws = wb.active
                self.excelData = [i[0] for i in list(ws.values)[1:]]
                self.totals = ws.max_row - 1
                showinfo("提示信息", "共导入{}条数据".format(self.totals))
                self.statusNowText.configure(text=f"{0}/{self.totals}")
            except:
                showerror("错误信息", "请导入正确的Excel文件!")
        else:
            showerror("错误信息", "请导入文件!")

    async def __getContent(self, semaphore, phoneNo, shopId, opId):
        link = "https://newsale.chnl.zj.chinamobile.com/newsale-web/wechat/checkRuleAndTime"
        headers = {
            'Host': 'newsale.chnl.zj.chinamobile.com',
            'Connection': 'keep-alive',
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36 MicroMessenger/7.0.9.501 NetType/WIFI MiniProgramEnv/Windows WindowsWechat',
            'content-type': 'application/json',
            'Referer': 'https://servicewechat.com/wxf26e1816ed35f0fd/83/page-frame.html',
            'Accept-Encoding': 'gzip, deflate, br'
        }

        # formData = {"bill_id": phoneNo, "program_id": "600000661055", "kind_id": "600000661056",
        #             "orgId": shopId, "offerInfoParamList": [
        #         {"offer_id": "600000655085", "offer_type": "OFFER_VAS_PREPAY", "offer_oper_type": 1}]}
        formData = {"bill_id": phoneNo, "program_id": "600000661055", "kind_id": "600000661056",
                    "orgId": shopId, "opId": opId, "offerInfoParamList": [
                {"offer_id": "600000655085", "offer_type": "OFFER_VAS_PREPAY", "offer_oper_type": 1}]}

        async with semaphore:
            conn = aiohttp.TCPConnector(verify_ssl=False)
            async with aiohttp.ClientSession(connector=conn, headers=headers) as session:
                try:
                    async with await session.post(link, data=json.dumps(formData), timeout=3) as resp:
                        content = await resp.json()
                        await asyncio.sleep(self.sleepTime)

                        return content
                except:
                    return

    async def __crawler(self, semaphore, phoneNo, shopId, opId):
        try:
            content = await self.__getContent(semaphore, phoneNo, shopId, opId)
            try:
                result = []
                result01 = "是" if not content[0].get("RESP_PARAM").get("BUSI_INFO").get("CHECKRSLTLIST").get("CHECKRSLTINFO").get("ERRORLIST") else "否"
                result02 = "是" if not content[-1].get("RESP_PARAM").get("BUSI_INFO").get("OFFER_LIST").get("OFFER_INFO").get("ERR_LIST").get("ERR_INFO") else "否"

                result.append(result01)
                result.append(result02)
                print(result)
                if "否" in result:
                    result = False
            except Exception as e:
                print(e.args)
                print(content)
                result = False


            treeData = [
                self.treeIndex,
                phoneNo,
                "是" if result else "否"
            ]
            self.totalData.append(treeData[1:])
            self.box.insert("", "end", values=treeData)
            self.statusNowText.configure(text=f"{self.treeIndex}/{self.totals}")
            self.treeIndex += 1
            self.box.yview_moveto(1.0)
            self.excelData.remove(phoneNo)
        except Exception as e:
            print(e.args)
            pass

    def __saveExcel(self):
        if len(self.totalData) <= 1:
            showerror("错误信息", "当前不存在任何数据!")
            return

        excelPath = filedialog.asksaveasfilename(title=u'保存文件', filetypes=[("xlsx", ".xlsx")]) + ".xlsx"
        if excelPath.strip(".xlsx"):
            wb = Workbook()
            ws = wb.active
            for line in self.totalData:
                ws.append(line)
            wb.save(excelPath)
            showinfo("提示信息", "保存成功！")

    async def taskManager(self, dataList, shopId, func, opId):
        tasks = []
        sem = asyncio.Semaphore(self.semNum)
        for phone in dataList:
            task = asyncio.ensure_future(func(sem, phone, shopId, opId))
            tasks.append(task)
        await asyncio.gather(*tasks)

    def start(self):
        try:
            self.sleepTime = float(self.entrySleep.get().strip())
        except:
            self.sleepTime = 0.01
            showinfo("提示", "延时大小有误,使用默认延时0.01秒!")

        shopId = self.shopIdText.get().strip()
        if re.findall("\D+", shopId):
            showerror("错误信息", "请输入正确的店铺ID!")
            return

        opId = self.workNoText.get().strip()

        if self.totals > 0:
            new_loop = asyncio.new_event_loop()
            asyncio.set_event_loop(new_loop)
            self.loop = asyncio.get_event_loop()
            self.loop.run_until_complete(self.taskManager(self.excelData, shopId, self.__crawler, opId))
        else:
            showerror("错误信息", "请导入数据!")
            return

        while self.excelData:
            self.loop = asyncio.get_event_loop()
            self.loop.run_until_complete(self.taskManager(self.excelData, shopId, self.__crawler, opId))

        showinfo("提示信息", "任务结束！")

    def stop(self):
        for task in asyncio.Task.all_tasks(self.loop):
            task.cancel()
            self.loop.stop()
            self.loop.run_forever()
        self.loop.close()

    @staticmethod
    def thread_it(func, *args):
        myThred = threading.Thread(target=func, args=args)
        myThred.setDaemon(True)
        myThred.start()

    def run(self):
        self.root.mainloop()


if __name__ == '__main__':
    app = Application()
    app.run()
    """
    店铺ID  10203121
    工号 20398477
    """
